"""
Importe la nomenclature officielle des métiers de l'ANEM (Agence Nationale de
l'Emploi, Algérie) depuis un fichier Excel, en remplacement complet de l'ancien
référentiel (ROME français + scraping Emploitic).

Colonnes attendues dans le fichier (feuille unique) :
    Secteurs | CodeDomaine | Domaines | sousDomaines | CodeFicheName | FichesNames | Appelations

Remplace entièrement Secteur/Domaine/SousDomaine/MetierReferentiel — pas de fusion
avec les données existantes.

Usage :
    python manage.py import_anem_nomenclature --dry-run
    python manage.py import_anem_nomenclature
    python manage.py import_anem_nomenclature --migrate-existing-data
    python manage.py import_anem_nomenclature --file chemin/vers/fichier.xlsx
"""
import os
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

# Mapping best-effort : ancien code SECTEURS_CHOICES (pré-ANEM) -> code Domaine ANEM
# le plus proche. Utilisé uniquement pour migrer les quelques enregistrements
# existants (offres, profils) créés avant le passage à la nomenclature ANEM.
ANCIEN_VERS_DOMAINE = {
    'IT': 'L18',            # Systèmes d'information et de télécommunication
    'INGENIERIE': 'C1',     # Études et supports techniques à l'industrie
    'COMMERCIAL': 'L17',    # Stratégie commerciale et marketing
    'VENTE': 'G16',         # Force de vente
    'GRANDS_COMPTES': 'L17',
    'MARKETING': 'L17',
    'FINANCE': 'L12',       # Comptabilité et finance
    'LOGISTIQUE': 'I1',     # Logistique
    'PRODUCTION': 'C2',     # Production industrielle
    'BTP': 'F11',           # Direction et conduite de chantier en BTP
    'ADMIN': 'L16',         # Secrétariat et assistance
    'SECRETARIAT': 'L16',
    'RH': 'L15',            # Ressources humaines
    'SANTE': 'M11',         # Praticiens médicaux
    'QSE': 'C1',            # Pas d'équivalent QSE direct — repli sur études techniques industrie
    'TOURISME': 'H11',      # Gestion et direction établissement hôtelier/restauration
    'MAINTENANCE': 'D1',    # Installation et maintenance
    'JURIDIQUE': 'P16',     # Droit
    'AUTRE': None,
}


class Command(BaseCommand):
    help = "Importe la nomenclature ANEM (Secteur/Domaine/SousDomaine/MetierReferentiel) depuis un fichier Excel."

    def add_arguments(self, parser):
        parser.add_argument("--dry-run", action="store_true")
        parser.add_argument("--migrate-existing-data", action="store_true",
                             help="Remappe les OffreEmploi/ProfilCandidat/ExperienceCandidat existants vers un code Domaine (via ANCIEN_VERS_DOMAINE) au lieu d'importer la nomenclature.")
        parser.add_argument("--file", type=str, default=None,
                             help="Chemin du fichier Excel ANEM (défaut : NAME.xlsx à la racine du projet).")

    def handle(self, *args, **options):
        if options["migrate_existing_data"]:
            return self._migrer_donnees_existantes(options["dry_run"])
        return self._importer_nomenclature(options["dry_run"], options["file"])

    # ------------------------------------------------------------------
    def _importer_nomenclature(self, dry_run, file_path):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write("openpyxl requis : pip install openpyxl")
            return

        path = file_path or os.path.join(settings.BASE_DIR, "..", "NAME.xlsx")
        path = os.path.abspath(path)
        if not os.path.exists(path):
            self.stderr.write(f"Fichier introuvable : {path}")
            return

        self.stdout.write(f"Lecture de {path}...")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb[wb.sheetnames[0]]

        secteurs = {}     # code -> libelle
        domaines = {}     # code -> (libelle, secteur_code)
        sous_domaines = set()  # (domaine_code, libelle)
        appellations = []  # liste de dicts, dédupliquée sur (code_fiche, titre)
        vus = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            secteur_brut, code_dom, domaine_lbl, sous_dom, code_fiche, fiche_lbl, appellation = [
                v.strip() if isinstance(v, str) else v for v in row
            ]
            if not (code_dom and appellation and code_fiche):
                continue

            secteur_code, _, secteur_lbl = secteur_brut.partition('-')
            secteur_code = secteur_code.strip()
            secteur_lbl = secteur_lbl.strip().title()
            secteurs.setdefault(secteur_code, secteur_lbl)

            domaines.setdefault(code_dom, (domaine_lbl, secteur_code))

            sous_dom_val = sous_dom if sous_dom and sous_dom != 'NULL' else None
            if sous_dom_val:
                sous_domaines.add((code_dom, sous_dom_val))

            cle = (code_fiche, appellation)
            if cle in vus:
                continue
            vus.add(cle)
            appellations.append({
                "titre": appellation,
                "domaine_code": code_dom,
                "sous_domaine": sous_dom_val,
                "code_fiche": code_fiche,
                "fiche_metier": fiche_lbl or "",
                "secteur_code": secteur_code,
            })

        self.stdout.write(
            f"Détecté : {len(secteurs)} secteurs, {len(domaines)} domaines, "
            f"{len(sous_domaines)} sous-domaines, {len(appellations)} appellations."
        )

        if dry_run:
            self.stdout.write("\n--- Aperçu secteurs ---")
            for code, lbl in list(secteurs.items())[:16]:
                self.stdout.write(f"  {code} — {lbl}")
            self.stdout.write("\n--- Aperçu appellations ---")
            for a in appellations[:10]:
                self.stdout.write(f"  {a['titre']} ({a['domaine_code']} / {a['code_fiche']})")
            self.stdout.write("\n[dry-run] Aucune écriture effectuée.")
            return

        from ...models import Secteur, Domaine, SousDomaine, MetierReferentiel

        with transaction.atomic():
            MetierReferentiel.objects.all().delete()
            SousDomaine.objects.all().delete()
            Domaine.objects.all().delete()
            Secteur.objects.all().delete()

            Secteur.objects.bulk_create([
                Secteur(code=code, libelle=lbl) for code, lbl in secteurs.items()
            ], ignore_conflicts=True)
            secteur_par_code = {s.code: s for s in Secteur.objects.all()}

            Domaine.objects.bulk_create([
                Domaine(code=code, libelle=lbl, secteur=secteur_par_code[sect_code])
                for code, (lbl, sect_code) in domaines.items()
                if sect_code in secteur_par_code
            ], ignore_conflicts=True)
            domaine_par_code = {d.code: d for d in Domaine.objects.all()}

            sous_domaine_objs = SousDomaine.objects.bulk_create([
                SousDomaine(domaine=domaine_par_code[dom_code], libelle=lbl)
                for dom_code, lbl in sous_domaines
                if dom_code in domaine_par_code
            ], ignore_conflicts=True)
            sous_domaine_par_cle = {
                (sd.domaine_id, sd.libelle): sd for sd in SousDomaine.objects.select_related('domaine')
            }

            metiers = []
            for a in appellations:
                domaine = domaine_par_code.get(a["domaine_code"])
                if not domaine:
                    continue
                sous_domaine = None
                if a["sous_domaine"]:
                    sous_domaine = sous_domaine_par_cle.get((domaine.id, a["sous_domaine"]))
                metiers.append(MetierReferentiel(
                    titre=a["titre"],
                    domaine=domaine,
                    sous_domaine=sous_domaine,
                    code_fiche=a["code_fiche"],
                    fiche_metier=a["fiche_metier"],
                    secteur_code=a["secteur_code"],
                ))
            MetierReferentiel.objects.bulk_create(metiers, ignore_conflicts=True, batch_size=500)

        self.stdout.write(self.style.SUCCESS(
            f"OK — {Secteur.objects.count()} secteurs, {Domaine.objects.count()} domaines, "
            f"{SousDomaine.objects.count()} sous-domaines, {MetierReferentiel.objects.count()} métiers importés."
        ))

    # ------------------------------------------------------------------
    def _migrer_donnees_existantes(self, dry_run):
        from ...models import OffreEmploi, ProfilCandidat, ExperienceCandidat

        cibles = [
            (OffreEmploi, 'specialite'),
            (ProfilCandidat, 'specialite'),
            (ProfilCandidat, 'secteur_souhaite'),
            (ExperienceCandidat, 'secteur'),
        ]
        total = 0
        for model, champ in cibles:
            qs = model.objects.filter(**{f"{champ}__in": list(ANCIEN_VERS_DOMAINE.keys())})
            self.stdout.write(f"{model.__name__}.{champ} : {qs.count()} enregistrement(s) à remapper.")
            if dry_run:
                continue
            for obj in qs:
                ancien = getattr(obj, champ)
                nouveau = ANCIEN_VERS_DOMAINE.get(ancien)
                setattr(obj, champ, nouveau)
                obj.save(update_fields=[champ])
                total += 1

        if dry_run:
            self.stdout.write("[dry-run] Aucune écriture effectuée.")
        else:
            self.stdout.write(self.style.SUCCESS(f"OK — {total} enregistrement(s) remappés vers un code Domaine ANEM."))
